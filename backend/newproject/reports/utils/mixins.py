from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from weasyprint import HTML
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

class ReportExportMixin:
    def get_template_name(self):
        raise NotImplementedError("Subclasses must implement get_template_name()")
    
    def get_export_filename(self, export_format, data):
        report_name = self.__class__.__name__.replace('View', '').replace('Report', '')
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        extension = 'pdf' if export_format == 'pdf' else 'xlsx'
        return f"{report_name}_{timestamp}.{extension}"
    
    def generate_pdf_response(self, data):
        context = {
            'current_date': timezone.now(),
            'data': data
        }
        
        template_name = self.get_template_name()
        html_string = render_to_string(template_name, context)
        html = HTML(string=html_string)
        
        filename = self.get_export_filename('pdf', data)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        pdf = html.write_pdf()
        response.write(pdf)
        return response
    
    def generate_excel_response(self, data):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create sheets based on data structure
        self.create_excel_sheets(wb, data)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = self.get_export_filename('excel', data)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def create_excel_sheets(self, wb, data):

        for key, value in data.items():
            if isinstance(value, list) and value:  
                sheet = wb.create_sheet(key.replace('_', ' ').title())
                self.populate_sheet_with_list_data(sheet, value, key)
    
    def populate_sheet_with_list_data(self, sheet, data_list, data_key):
        if not data_list:
            return

        if isinstance(data_list[0], dict):
            headers = list(data_list[0].keys())
        else:
            headers = ["Value"]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="223152", end_color="223152", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
        
        # Populate data
        for row_idx, item in enumerate(data_list, 2):
            if isinstance(item, dict):
                for col_idx, header in enumerate(headers, 1):
                    value = item.get(header, '')
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            else:
                sheet.cell(row=row_idx, column=1, value=item)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
